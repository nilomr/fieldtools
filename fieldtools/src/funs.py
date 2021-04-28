import datetime
import glob
import inspect
import os
import re
import shutil
import sys
import time
import warnings
from getpass import getuser
from subprocess import PIPE, Popen, check_output

import pandas as pd
import psutil
import pygsheets
from fieldtools.src.aesthetics import arrow, info, tcolor, tstyle
from fieldtools.src.paths import OUT_DIR, PROJECT_DIR, safe_makedir
from openpyxl.reader.excel import load_workbook
from pathlib2 import Path, PosixPath
from tqdm.auto import tqdm

warnings.simplefilter(action='ignore', category=UserWarning)


def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=None,
    truncate_sheet=False,
    **to_excel_kwargs
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl")

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def safe_makedir(FILE_DIR):
    """Make a safely nested directory.
    Args:
        FILE_DIR (str or PosixPath): Path to be created.
    """
    if type(FILE_DIR) == str:
        if "." in os.path.basename(os.path.normpath(FILE_DIR)):
            directory = os.path.dirname(FILE_DIR)
        else:
            directory = os.path.normpath(FILE_DIR)
        if not os.path.exists(directory):
            try:
                os.makedirs(directory)
            except FileExistsError as e:
                # multiprocessing can cause directory creation problems
                print(e)
    elif type(FILE_DIR) == PosixPath:
        # if this is a file
        if len(FILE_DIR.suffix) > 0:
            FILE_DIR.parent.mkdir(parents=True, exist_ok=True)
        else:
            try:
                FILE_DIR.mkdir(parents=True, exist_ok=True)
            except OSError as e:
                if e.errno != 17:
                    print("Error:", e)


def write_gpx(filename, newboxes, tocollect):
    """Writes .gpx file to disk, containing 
    a) all great tit nestboxes that haven't been recorded (in green),
    b) nestboxes where recorders need to be collected (in red) and
    c) nestboxes that haven't been recored and have eggs (helipad symbol).

    Args:
        filename (PosixPath): path including filename and extension (.gpx) where to output file.
        newboxes (DataFrame): DataFrame containing all new boxes, with ['Nestbox', 'x', 'y'] columns.
        tocollect (DataFrame): DataFrame containing boxes from n days ago,  with ['Nestbox', 'x', 'y'] columns.
    """
    safe_makedir(filename)
    gpxfile = open(str(filename), "w")
    gpxfile.write(
        '<?xml version="1.0"?><gpx version="1.1" creator="Nilo M. Recalde" >'
    )

    try:
        allpoints = newboxes.query(
            'Eggs == "no"').filter(["Nestbox", "longitude", "latitude"])
        allpoints_transformed = allpoints.assign(
            **{"lon": allpoints["longitude"], "lat": allpoints["latitude"]}
        ).to_dict(orient="records")

        for box in allpoints_transformed:
            poi = '<wpt lat="{}" lon="{}"><name>{}</name><sym>{}</sym></wpt>'.format(
                box["lat"], box["lon"], box["Nestbox"], "poi_green"
            )
            gpxfile.write(poi)

    except Exception:
        pass

    try:
        eggs = newboxes.query('Eggs != "no"').filter(
            ["Nestbox", "longitude", "latitude"])
        eggs_transformed = eggs.assign(
            **{"lon": eggs["longitude"], "lat": eggs["latitude"]}
        ).to_dict(orient="records")

        for box in eggs_transformed:
            poi = '<wpt lat="{}" lon="{}"><name>{}</name><sym>{}</sym></wpt>'.format(
                box["lat"], box["lon"], box["Nestbox"], "helipad"
            )
            gpxfile.write(poi)

    except Exception:
        pass

    try:
        collect = tocollect.filter(["Nestbox", "longitude", "latitude"])
        collect_transformed = collect.assign(
            **{"lon": collect["longitude"], "lat": collect["latitude"]}
        ).to_dict(orient="records")

        for box in collect_transformed:
            poi = '<wpt lat="{}" lon="{}"><name>{}</name><sym>{}</sym></wpt>'.format(
                box["lat"], box["lon"], box["Nestbox"], "poi_red"
            )
            gpxfile.write(poi)

    except Exception:
        pass

    gpxfile.write("</gpx>")
    gpxfile.close()


def order(frame, var):
    if type(var) is str:
        var = [var]  # let the command take a string or list
    varlist = [w for w in frame.columns if w not in var]
    frame = frame[var + varlist]
    return frame


def get_faceplate_update():
    gc = pygsheets.authorize(
        service_file=str(PROJECT_DIR / "private" / "client_secret.json")
    )
    googlekey = 'ABCDEF'  # The faceplating sheet, substitute your own
    faceplate_info = (
        gc.open_by_key(googlekey)[0]
        .get_as_df(has_header=True, include_tailing_empty=False)
        .filter(['Nestbox', 'Species'])
        .query('Species == "g" or Species == "G" or Species == "sp=g"')
    )
    print(arrow + "Downloading species data: 100%")
    return faceplate_info['Nestbox'].tolist()


def get_comments_update():
    gc = pygsheets.authorize(
        service_file=str(PROJECT_DIR / "private" / "client_secret.json")
    )
    googlekey = 'ABCDEF'  # The comments sheet, substitute your own
    comments_df = (
        gc.open_by_key(googlekey)[0]
        .get_as_df(has_header=True, include_tailing_empty=False)
        .query('Again == "YES" or Again == "TRUE"')
        .filter(['Nestbox', 'Comments'])
    )
    return comments_df


class workers:
    """
    A dictionary of google sheet keys, provide your own - this is just a placeholder
    """

    gdict = {  # ! Change every year
        "Jane": "AB23CD3sry3F",
        "John": "AB23CD3sry3F"
    }
    rounds_dict = {
        'Bean': 'Jane',
        'Broad Oak': 'John',
    }


def get_nestbox_update():
    gc = pygsheets.authorize(
        service_file=str(PROJECT_DIR / "private" / "client_secret.json")
    )
    # Download and append personal sheets
    workerdict = workers.gdict
    which_greti = pd.DataFrame(columns=["Nestbox", "Owner"])

    for worker, googlekey in tqdm(
        workerdict.items(),
        desc=arrow + "Downloading field worker data",
        position=0,
        leave=True,
        bar_format='{desc}: {percentage:3.0f}%'
    ):
        worker = gc.open_by_key(googlekey)[0].get_as_df(
            has_header=False, include_tailing_empty=False)
        worker = worker.rename(
            columns=worker.iloc[0]).drop(worker.index[0])
        if "" in worker.columns:
            worker = worker.drop([""], axis=1)
        worker = (
            worker.query("Nestbox == Nestbox")
            .rename(columns={"Clutch Size": "Eggs", "Num eggs": "Eggs", "Num Eggs": "Eggs", "Clutch size": "Eggs", "State code": "Nest", "State Code": "Nest"})
            .filter(["Nestbox", "Owner", "Eggs", 'Nest', 'Species'])
            .replace("", "no")
        )
        which_greti = which_greti.append(worker)

    # Now get faceplating info and join
    greti_faceplated = get_faceplate_update()
    combined = which_greti.query(
        'Species == "g" or Species == "G" or Species == "sp=g" or Nestbox in @greti_faceplated').drop('Species', 1)

    return combined


def get_single_gsheet(name, key):
    gc = pygsheets.authorize(
        service_file=str(PROJECT_DIR / "private" /
                         "client_secret.json")
    )
    worker = gc.open_by_key(key)[0].get_as_df(
        has_header=False, include_tailing_empty=False)
    worker = worker.rename(
        columns=worker.iloc[0]).drop(worker.index[0])
    if "" in worker.columns:
        worker = worker.drop([""], axis=1)
    worker = (
        worker.query("Nestbox == Nestbox")
        .rename(columns={"Clutch Size": "Eggs", "Num eggs": "Eggs", "Num Eggs": "Eggs", "Clutch size": "Eggs", "State code": "Nest", "State Code": "Nest"})
        .filter(["Nestbox", "Owner", "Eggs", 'Nest', 'Species'])
        .replace("", "no")
    )
    return worker.query("Nestbox != 'no'")


def get_recorded_gretis(recorded_csv, nestbox_coords, which_greti):
    picklename = OUT_DIR / (
        str(
            f"allrounds_{str(pd.Timestamp('today', tz='UTC').strftime('%Y%m%d'))}.pkl"
        )
    )
    if len(which_greti) == 0:
        print(info + "There are no GRETI nestboxes yet")
        return [], []
    else:
        which_greti = pd.merge(
            which_greti, nestbox_coords, on=["Nestbox"])
        which_greti["Added"] = str(
            pd.Timestamp("today", tz="UTC").strftime("%Y-%m-%d")
        )
        len1 = len(which_greti)
        # Remove Blue tit nestboxes from list
        greti_boxes = nestbox_coords.query(
            '`box type` == "GT"')['Nestbox'].to_list()
        bluti_boxes = nestbox_coords.query(
            '`box type` == "BT"')['Nestbox'].to_list()
        which_greti_1 = which_greti[which_greti['Nestbox'].isin(greti_boxes)]
        which_wrong = which_greti[which_greti['Nestbox'].isin(bluti_boxes)]
        len2 = len(which_greti_1)
        if len1 != len2:
            print(
                info + f'Removed {len1 - len2} nestboxes that were of blue tit type')
            # print(which_wrong)
        which_greti_1.to_pickle(str(picklename))

    # Check which nestboxes have already been recorded
    if not Path(recorded_csv).exists():
        print(info + '.csv file with recorded nestboxes does not exist, creating it.')
        recorded_empty = pd.DataFrame(
            columns=['Nestbox', 'AM', 'longitude', 'latitude', 'Deployed', 'Move_by'])
        recorded_empty.to_csv(recorded_csv, index=False)
    try:
        already_recorded = (
            pd.read_csv(recorded_csv)
            .filter(["Nestbox"])
            .query('Nestbox != "Nestbox"')
        )
        diff_df = (
            which_greti_1.merge(
                already_recorded, on=["Nestbox"], indicator=True, how="outer"
            )
            .query('_merge != "both"')
            .drop(["_merge"], 1)
            .dropna(thresh=2)
        )
        if {"longitude_x", "latitude_y"}.issubset(diff_df.columns):
            diff_df = diff_df.drop(["longitude_y", "latitude_y"], 1).rename(
                columns={"longitude_x": "longitude",
                         "latitude_x": "latitude"}
            )
        diff_df['Nest'] = diff_df['Nest'].replace('no', 0)
        diff_df = diff_df.sort_values(
            ['Eggs', 'Nest'], ascending=[True, False])

    except:
        already_recorded = []
        diff_df = which_greti_1
    try:
        comments = get_comments_update()
        diff_df = pd.merge(
            diff_df, comments, how="left", on=["Nestbox"])
        diff_df.fillna('', inplace=True)
    except:
        print('Error when downloading comments, skipping comments.')

    return already_recorded, diff_df


def split_path(path):
    folders = []
    drive, path = os.path.splitdrive(path)
    while True:
        path, folder = os.path.split(path)

        if folder:
            folders.append(folder)
        else:
            if path:
                folders.append(path)
            break

    if drive:
        folders.append(drive)
    return folders[::-1]


def reconstruct_path(folders):
    folders = folders[:]
    path = ""

    # On windows, pop off the drive if there is one. Affects handling of relative vs rooted paths
    if sys.platform == 'win32' and ':' == folders[0][-1]:
        path = folders[0]
        del folders[0]
    if folders and folders[0] == os.sep:
        path += folders[0]
        del folders[0]
    path += os.sep.join(folders)
    return path


# Format cards

def get_mountedlist():
    return [card[card.find("/"):] for card in check_output(
            ["/bin/bash", "-c", "lsblk"]).decode("utf-8").split("\n") if "/" in card]


def is_faceplate(dev):
    try:
        nm = Path(dev).name
        if nm[0] == 'F' and len(nm) == 5 and nm[1:4].isnumeric():
            return True
        else:
            return False
    except:
        return False


def yes_or_no(question):
    while "The answer is invalid":
        reply = str(input(question + " [y/n]: ")).lower().strip()
        if reply[0] == "y":
            return True
        elif reply[0] == "n":
            return False
        else:
            print(tcolor("The answer is invalid", tstyle.rojoroto))


def find_sdiskpart(path):
    path = os.path.abspath(path)
    while not os.path.ismount(path):
        path = os.path.dirname(path)
    p = [p for p in psutil.disk_partitions(
        all=True) if p.mountpoint == path.__str__()]
    l = len(p)
    if len(p) == 1:
        return p[0]
    raise psutil.Error


def get_wav_filenames(card):
    return [os.sep.join(os.path.normpath(file).split(os.sep)[-2:])
            for file in [os.path.join(card[0], i)
                         for i in os.listdir(card[0])
                         if i.endswith('.WAV')]
            ]


def parse_blkid(valid_directories, output):
    valid_devices = []
    output = [str(i).split(' ') for i in output.split(b'\n') if b'LABEL' in i]
    for ls in output:
        for sbls in ls:
            if sbls.startswith('LABEL'):
                name = re.findall(r'"(.*?)"', sbls)[0]
                if (
                    name[0] == 'F' and len(name) == 5 and
                    name[1:4].isnumeric() or
                    name in [am[0] for am in valid_directories]
                ):
                    devpath = str(ls[0]).replace(
                        ':', '').replace('b\'', '')
                    valid_devices.append([name, devpath])
    return valid_devices


def ensure_mount(valid_directories, password, checked_cards, already_done, verbose):
    devnull = open(os.devnull, 'wb')
    mounted = get_mountedlist()
    blkid = "sudo blkid"
    proc1 = Popen(["/bin/bash", "-c", blkid],
                  stdin=PIPE, stdout=PIPE)
    output, err = proc1.communicate()
    valid_devices = parse_blkid(valid_directories, output)
    if not valid_devices:
        pass
    else:
        for dev in valid_devices:
            if any(dev[0] in s for s in mounted):
                continue
            elif dev[0] in already_done:
                continue
            else:

                # Mkdir
                target_dir = os.path.join(os.sep, 'media', getuser(), dev[0])

                if os.path.exists(target_dir):
                    if verbose:
                        print(
                            tcolor(f'The mount point {target_dir} already exists', tstyle.rojoroto))
                    # Delete mount point
                    delmount = f"sudo rmdir {target_dir}"
                    proc5 = Popen(["/bin/bash", "-c", delmount],
                                  stdin=PIPE, stdout=PIPE, stderr=devnull)
                    proc5.communicate()

                mkdir = f"sudo mkdir -p {target_dir}"
                proc2 = Popen(["/bin/bash", "-c", mkdir],
                              stdin=PIPE, stdout=PIPE, stderr=devnull)
                out2, err2 = proc2.communicate()
                # Mount
                dmount = f"sudo mount {dev[1]} {target_dir}"
                proc3 = Popen(["/bin/bash", "-c", dmount],
                              stdin=PIPE, stdout=PIPE, stderr=devnull)
                out3, err3 = proc3.communicate()
                checked_cards.append(dev[0])
    return checked_cards


def umount_and_rmdir(password, card):
    devnull = open(os.devnull, 'wb')
    umount = f"sudo umount {card[0]}"
    proc4 = Popen(["/bin/bash", "-c", umount],
                  stdin=PIPE, stdout=PIPE, stderr=devnull)
    proc4.communicate()
    # Delete mount point
    delmount = f"sudo rmdir {card[0]}"
    proc5 = Popen(["/bin/bash", "-c", delmount],
                  stdin=PIPE, stdout=PIPE, stderr=devnull)
    proc5.communicate()
    # print(f'delete mountpoint returned {proc5.returncode}')


def any_mounted(mountdir):
    for nm in os.listdir(mountdir):
        if nm == 'F' and len(nm) == 5 and nm[1:4].isnumeric():
            return True
        elif nm.startswith('AM'):
            return True
        else:
            return False


def clean_vols():
    devnull = open(os.devnull, 'wb')
    print(info + 'Cleaning mounted volumes')
    mountdir = os.path.join(os.sep, 'media', getuser()) + os.sep
    print(any_mounted(mountdir))
    while any_mounted(mountdir):
        print(
            info + tcolor('Please remove any cards from the card reader', tstyle.rojoroto), end="\r")
        time.sleep(1)

        umount = f"sudo umount {mountdir}F* {mountdir}AM*"
        proc4 = Popen(["/bin/bash", "-c", umount],
                      stdin=PIPE, stdout=PIPE, stderr=devnull)
        proc4.communicate()

        # Delete mount point
        delmount = f"sudo rmdir {mountdir}F* {mountdir}AM*"
        proc5 = Popen(["/bin/bash", "-c", delmount],
                      stdin=PIPE, stdout=PIPE, stderr=devnull)
        proc5.communicate()

# Copy cards


def progress_percentage(perc, width=None):
    # By flutefreak7,
    # https://stackoverflow.com/a/48450305
    # This will only work for python 3.3+ due to use of
    # os.get_terminal_size the print function etc.

    FULL_BLOCK = '█'
    # this is a gradient of incompleteness
    INCOMPLETE_BLOCK_GRAD = ['░', '▒', '▓']

    assert(isinstance(perc, float))
    assert(0. <= perc <= 100.)
    # if width unset use full terminal
    if width is None:
        width = os.get_terminal_size().columns
    # progress bar is block_widget separator perc_widget : ####### 30%
    max_perc_widget = '100.00%'  # 100% is max
    separator = ' '
    blocks_widget_width = width - len(separator) - len(max_perc_widget)
    assert(blocks_widget_width >= 10)  # not very meaningful if not
    perc_per_block = 100.0/blocks_widget_width
    # epsilon is the sensitivity of rendering a gradient block
    epsilon = 1e-6
    # number of blocks that should be represented as complete
    full_blocks = int((perc + epsilon)/perc_per_block)
    # the rest are "incomplete"
    empty_blocks = blocks_widget_width - full_blocks

    # build blocks widget
    blocks_widget = ([FULL_BLOCK] * full_blocks)
    blocks_widget.extend([INCOMPLETE_BLOCK_GRAD[0]] * empty_blocks)
    # marginal case - remainder due to how granular our blocks are
    remainder = perc - full_blocks*perc_per_block
    # epsilon needed for rounding errors (check would be != 0.)
    # based on reminder modify first empty block shading
    # depending on remainder
    if remainder > epsilon:
        grad_index = int((len(INCOMPLETE_BLOCK_GRAD)
                          * remainder)/perc_per_block)
        blocks_widget[full_blocks] = INCOMPLETE_BLOCK_GRAD[grad_index]

    # build perc widget
    str_perc = '%.2f' % perc
    # -1 because the percentage sign is not included
    perc_widget = '%s%%' % str_perc.ljust(len(max_perc_widget) - 1)

    # form progressbar
    progress_bar = '%s%s%s' % (''.join(blocks_widget), separator, perc_widget)
    # return progressbar as string
    return ''.join(progress_bar)


def copy_progress(copied, total):
    print('\r' + progress_percentage(100*copied/total, width=30), end='')


def fetch_recorder_info(recorders_dir):
    try:
        recorders_info = pd.read_csv(
            recorders_dir).query('Nestbox != "Nestbox"')
    except:
        raise FileNotFoundError
    # Stop if no info on file
    if len(recorders_info) == 0:
        raise IndexError

    recorders_info['Deployed'] = pd.to_datetime(
        recorders_info['Deployed'], format='%Y-%m-%d')
    recorders_info['Move_by'] = pd.to_datetime(
        recorders_info['Move_by'], format='%Y-%m-%d')

    return recorders_info


def copyfile(src, dst, *, follow_symlinks=True):
    """Copy data from src to dst.

    If follow_symlinks is not set and src is a symbolic link, a new
    symlink will be created instead of copying the file it points to.

    """
    # By flutefreak7,
    # https://stackoverflow.com/a/48450305
    if shutil._samefile(src, dst):
        raise shutil.SameFileError(
            "{!r} and {!r} are the same file".format(src, dst))

    for fn in [src, dst]:
        try:
            st = os.stat(fn)
        except OSError:
            # File most likely does not exist
            pass
        else:
            # XXX What about other special files? (sockets, devices...)
            if shutil.stat.S_ISFIFO(st.st_mode):
                raise shutil.SpecialFileError("`%s` is a named pipe" % fn)

    if not follow_symlinks and os.path.islink(src):
        os.symlink(os.readlink(src), dst)
    else:
        size = os.stat(src).st_size
        with open(src, 'rb') as fsrc:
            with open(dst, 'wb') as fdst:
                copyfileobj(fsrc, fdst, callback=copy_progress, total=size)
    return dst


def copyfileobj(fsrc, fdst, callback, total, length=16*1024):
    copied = 0
    while True:
        buf = fsrc.read(length)
        if not buf:
            break
        fdst.write(buf)
        copied += len(buf)
        callback(copied, total=total)


def copy_with_progress(src, dst, *, follow_symlinks=True):

    if type(dst) == PosixPath:
        dst = str(dst)
    if os.path.isdir(dst):
        dst = os.path.join(dst, os.path.basename(src))
        print(f'\n{Path(dst).name}')

    copyfile(src, dst, follow_symlinks=follow_symlinks)
    shutil.copymode(src, dst)
    return dst


def get_nestbox_id(recorders_dir, recorders_info, card, am, filedate):
    try:
        nestbox = recorders_info[(recorders_info['AM'] == str(am)) | (
            recorders_info['AM'] == int(am)) | (recorders_info['AM'] == '0' + str(am))]
        nestbox = nestbox[(nestbox['Deployed'] + datetime.timedelta(hours=10) < filedate) &
                          (nestbox['Move_by'] + datetime.timedelta(hours=10) >= filedate)]['Nestbox']
    except:
        print(tcolor('\n\n' + inspect.cleandoc(f"""
                Unknown error when trying to get the recorder information 
                for file with datetime {filedate} from {card[1]}"""), tstyle.rojoroto))

    if len(nestbox) == 1:
        nestbox = nestbox.iat[0]
        return nestbox
    elif len(nestbox) > 1:
        print(tcolor('\n\n' + inspect.cleandoc(f"""
                There are more than one row compatible with this
                AM / date combination ({card[1]}, {filedate})"""), tstyle.rojoroto))
    else:
        print(tcolor('\n\n' + inspect.cleandoc(f"""
                There are no rows compatible with this
                AM / date combination ({card[1]}, {filedate}).
                Check that you have entered the deployment information in
                {recorders_dir}"""), tstyle.rojoroto))


def get_full_faceplate_info():
    gc = pygsheets.authorize(
        service_file=str(PROJECT_DIR / "private" /
                         "client_secret.json")
    )
    facekey = '1NToFktrKMan-jlGYnASMM_AXSv1gwG2dYqjTGCY-6lw'  # The faceplating sheet
    faceplate_info = (
        gc.open_by_key(facekey)[0]
        .get_as_df(has_header=True, include_tailing_empty=False))
    return faceplate_info
